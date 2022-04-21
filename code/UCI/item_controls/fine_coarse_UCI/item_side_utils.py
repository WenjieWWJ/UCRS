import math
import openpyxl
import numpy as np
import pandas as pd
from  openpyxl import load_workbook

def computeTopNAccuracy(GroundTruth, predictedIndices, topN, GroundTruthTarget=None):
    precision = [] 
    recall = [] 
    NDCG = [] 
    MRR = []
    WNDCG = []
    WNDCG_return = False
    
    if GroundTruthTarget is None:
        GroundTruthTarget = [[] for i in GroundTruth]
    else:
        WNDCG_return = True
    
    for index in range(len(topN)):
        sumForPrecision = 0
        sumForRecall = 0
        sumForNdcg = 0
        sumForMRR = 0
        sumForWndcg = 0
        
        for i in range(len(predictedIndices)):  # for a user,
            if len(GroundTruth[i]) != 0:
                mrrFlag = True
                userHit = 0
                userMRR = 0
                
                dcg = 0
                idcg = 0
                idcgCount = len(GroundTruth[i])
                ndcg = 0
                
                wdcg = 0
                widcg = 0
                widcgCount = len(GroundTruthTarget[i])
                wndcg = 0
                
                hit = []
                for j in range(topN[index]):
                    if predictedIndices[i][j] in GroundTruth[i]:
                        # if Hit!
                        dcg += 1.0/math.log2(j + 2)
                        if predictedIndices[i][j] in GroundTruthTarget[i]:
                            wdcg += 2.0/math.log2(j + 2)
                        else:
                            wdcg += 1.0/math.log2(j + 2)
                            
                        if mrrFlag:
                            userMRR = (1.0/(j+1.0))
                            mrrFlag = False
                        userHit += 1.0
                
                    if idcgCount > 0:
                        idcg += 1.0/math.log2(j + 2)
                        idcgCount = idcgCount-1
                        if widcgCount > 0:
                            widcg += 2.0/math.log2(j + 2)
                            widcgCount = widcgCount-1
                        else:
                            widcg += 1.0/math.log2(j + 2)
                            
                if(idcg != 0):
                    ndcg += (dcg/idcg)
                if widcg > 0:
                    wndcg += (wdcg/widcg)
                    
                sumForPrecision += userHit / topN[index]
                sumForRecall += userHit / len(GroundTruth[i])               
                sumForNdcg += ndcg
                sumForMRR += userMRR
                sumForWndcg += wndcg
        
        precision.append(round(sumForPrecision / len(predictedIndices), 4))
        recall.append(round(sumForRecall / len(predictedIndices), 4))
        NDCG.append(round(sumForNdcg / len(predictedIndices), 4))
        MRR.append(round(sumForMRR / len(predictedIndices), 4))
        WNDCG.append(round(sumForWndcg / len(predictedIndices), 4))
        
    if WNDCG_return:
        return precision, recall, NDCG, MRR, WNDCG
    return precision, recall, NDCG, MRR

def get_group_distribution(user_list, interaction_dict, item_feature, category_len, is_category_avg = True):
    distribution = [0] * category_len
    distribution_user = [0] * category_len
    for user in user_list:
        distribution = [0] * category_len
        for item in interaction_dict[user]:
            for cate in item_feature[item]:
                if is_category_avg == True:
                    distribution[cate] += 1/len(item_feature[item])
                else:
                    distribution[cate] += 1
        distribution_user = [distribution_user[i] + distribution[i]/len(interaction_dict[user]) for i in range(category_len)]
    distribution_avg = [i/len(user_list) for i in distribution_user]
    
    return distribution_avg

def get_distribution_df(feature_index, feature_group, training_dict, item_feature, second_class_list):
    df_list = []
    for cate in feature_group[feature_index].keys():
        distribution = get_group_distribution(feature_group[feature_index][cate], training_dict, item_feature, len(second_class_list), is_category_avg=False)
        df_list.append(distribution)

    distribution_df = pd.DataFrame(df_list)# .sort_values(by=[0], axis = 1, ascending=False)
#     # Convert category name into Chinese
#     for i in range(len(second_class_list)):
#         distribution_df.rename(columns={i: str(i) + '_' +id_second_class_map[i]},inplace=True)
    for i in range(len(feature_group[feature_index].keys())):
        
        index = list(feature_group[feature_index].keys())[i]
        distribution_df.loc[i, 'num'] = len(feature_group[feature_index][index])
        distribution_df.rename(index={i: str(index) + '_' + str(len(feature_group[feature_index][index]))},inplace=True)
    
    distribution_df.sort_values('num', ascending = False, inplace = True)
    distribution_df.drop('num', axis = 1, inplace = True)

    return distribution_df

def get_interest_group(train_dict, item_category, category_list, threshold_favorite_proportion, threshold_group_number):

    interest_group = {}

    for user in train_dict:
        distribution = get_group_distribution([user], train_dict, item_category, len(category_list), is_category_avg=False)
        if max(distribution) > threshold_favorite_proportion:
            interest = distribution.index(max(distribution))
            if interest not in interest_group:
                interest_group[interest] = [user]
            else:
                interest_group[interest].append(user)
                
    drop_key = []
    for key in interest_group:
        if len(interest_group[key]) < threshold_group_number:
            drop_key.append(key)
            
    for key in drop_key:
        interest_group.pop(key)
    
    interest_group_size = {}
    for interest in interest_group:
        interest_group_size[interest] = len(interest_group[interest])
    
    interest_group_size = dict(sorted(interest_group_size.items(), key=lambda e: e[1], reverse = True))
    sorted_interest_group = {}
    for interest in interest_group_size:
        sorted_interest_group[interest] = interest_group[interest]

                
    return sorted_interest_group

def get_GA_GD(train_dict, pred_dict, item_category, category_list, top_K_category, threshold, user_target):
    extreme_user_count = 0
    GD = {}
    GA = {}
    target_GA_avg = 0
    target_GD_avg = 0
    target_extreme_GA_avg = 0
    target_extreme_GD_avg = 0
    GA_all_avg = 0
    GA_extreme_avg = 0
    GD_all_avg = 0
    GD_extreme_avg = 0
    for user in pred_dict:
        
        amplification = 0
        domination = 0
        train_distribution = get_group_distribution([user], train_dict, item_category, len(category_list), is_category_avg=False)
        pred_distribution = get_group_distribution([user], pred_dict, item_category, len(category_list), is_category_avg=False)
        index_list=[i[0] for i in sorted(enumerate(train_distribution), key=lambda x:x[1])]

        for index in index_list[-top_K_category:][::-1]:
            amplification += pred_distribution[index]-train_distribution[index]
            domination += pred_distribution[index]

        GD[user] = domination    
        GA[user] = amplification # /top_K_category # do not average 
        
        ## target
        for cate in user_target[user]:
            target_GA_avg += pred_distribution[cate] - train_distribution[cate]
            target_GD_avg += pred_distribution[cate]

 
        if max(train_distribution) > threshold:
            extreme_user_count += 1
            GA_extreme_avg += amplification
            GD_extreme_avg += domination
            for cate in user_target[user]:
                target_extreme_GA_avg += pred_distribution[cate] - train_distribution[cate]
                target_extreme_GD_avg += pred_distribution[cate]

    GA_all_avg = sum(GA.values())/len(pred_dict)
    GA_extreme_avg = GA_extreme_avg/extreme_user_count
    GD_all_avg = sum(GD.values())/len(pred_dict)
    GD_extreme_avg = GD_extreme_avg/extreme_user_count
    
    target_GA_avg = target_GA_avg/len(pred_dict)
    target_GD_avg = target_GD_avg/len(pred_dict)
    target_extreme_GA_avg = target_extreme_GA_avg/extreme_user_count
    target_extreme_GD_avg = target_extreme_GD_avg/extreme_user_count


 
    
    all_user_num = len(pred_dict)
    
    # print('top_K_category = ', top_K_category)
    # print('threshold = ', threshold )
    # print('extreme user number = ',extreme_user_count)
    # print('average GA = ', sum(GA.values())/extreme_user_count)
    # print('average GD = ', sum(GD.values())/extreme_user_count)
    # print('-------------')
    return GA, GA_all_avg, GA_extreme_avg, GD, GD_all_avg, GD_extreme_avg, target_GA_avg, target_GD_avg, target_extreme_GA_avg, target_extreme_GD_avg, all_user_num, extreme_user_count

def get_GA_GD_all_threshold(threshold_proportion_list, top_K_category_list, train_dict, pred_dict, item_category, category_list, threshold_group_number, user_target):
    # top_K_category_list = [1, 2, 3, 4, 5, 10]
    interest_GA = {}
    interest_GD = {}
    GA_all_avg = {}
    GA_extreme_avg = {}
    GD_all_avg = {}
    GD_extreme_avg = {}
    all_user_num = {}
    extreme_user_num = {}

    target_GA_avg, target_GD_avg, target_extreme_GA_avg, target_extreme_GD_avg = {}, {}, {}, {}
    for threshold in threshold_proportion_list:
        interest_group = get_interest_group(train_dict, item_category, category_list, threshold, threshold_group_number)
        interest_GA[threshold] = {}
        interest_GD[threshold] = {}
        GA_all_avg[threshold] = {}
        GA_extreme_avg[threshold] = {}
        GD_all_avg[threshold] = {}
        GD_extreme_avg[threshold] = {}
        all_user_num[threshold] = {}
        extreme_user_num[threshold] = {}
        target_GA_avg[threshold], target_GD_avg[threshold], target_extreme_GA_avg[threshold], target_extreme_GD_avg[threshold] = {}, {}, {}, {}

        for top_K_category in top_K_category_list:
            GA, ga_all_avg, ga_extreme_avg, GD, gd_all_avg, gd_extreme_avg, target_ga_avg, target_gd_avg, target_extreme_ga_avg, target_extreme_gd_avg, all_user_num_temp, extreme_user_num_temp = get_GA_GD(train_dict, pred_dict, item_category, category_list, top_K_category, threshold, user_target)
            GA_all_avg[threshold][top_K_category] = ga_all_avg
            GA_extreme_avg[threshold][top_K_category] = ga_extreme_avg
            GD_all_avg[threshold][top_K_category] = gd_all_avg
            GD_extreme_avg[threshold][top_K_category] = gd_extreme_avg



            interest_GA[threshold][top_K_category] = {}
            interest_GD[threshold][top_K_category] = {}

            for interest in interest_group:
                interest_GA[threshold][top_K_category][interest] = 0
                interest_GD[threshold][top_K_category][interest] = 0
                for user in interest_group[interest]:
                    interest_GA[threshold][top_K_category][interest] += GA[user]
                    interest_GD[threshold][top_K_category][interest] += GD[user]
                interest_GA[threshold][top_K_category][interest] /= len(interest_group[interest])
                interest_GD[threshold][top_K_category][interest] /= len(interest_group[interest])

        all_user_num[threshold], extreme_user_num[threshold] = all_user_num_temp, extreme_user_num_temp
        target_GA_avg[threshold], target_GD_avg[threshold], target_extreme_GA_avg[threshold], target_extreme_GD_avg[threshold] = target_ga_avg, target_gd_avg, target_extreme_ga_avg, target_extreme_gd_avg
    return interest_GA, interest_GD, target_GA_avg, target_GD_avg, GA_all_avg, GD_all_avg, GA_extreme_avg, GD_extreme_avg, target_extreme_GA_avg, target_extreme_GD_avg, all_user_num, extreme_user_num

def get_category_num(user_dict, item_category, threshold):
        
    user_category_num = {}
    for user in user_dict:
        category_num = {}
        for item in user_dict[user]:
            for cate in item_category[item]:
                if cate not in category_num:
                    category_num[cate] = 1
                else:
                    category_num[cate] += 1
        user_category_num[user] = sum([1 for i in category_num.values() if i > threshold])
    return user_category_num
                   
                   
def get_category_A_D(train_dict, pred_dict, item_category, threshold):
    
    train_category = {}
    pred_category = {}
    category_A = {}
    category_D = {}
    train_category = get_category_num(train_dict, item_category, threshold)
    pred_category = get_category_num(pred_dict, item_category, threshold)
    
    for user in train_category:
        category_A[user] = pred_category[user] - train_category[user]
    category_D = pred_category   
    
    return category_A, category_D
    
def get_category_A_D_threshold(train_dict, pred_dict, item_category, threshold_category_list):
    category_A = {}
    category_D = {}
    category_A_avg = {}
    category_D_avg = {}
    
    for threshold in threshold_category_list:
        category_A[threshold], category_D[threshold] = get_category_A_D(train_dict, pred_dict, item_category, threshold)
        category_A_avg[threshold] = sum([i for i in category_A[threshold].values()])/len(category_A[threshold])
        category_D_avg[threshold] = sum([i for i in category_D[threshold].values()])/len(category_D[threshold])
    return category_A, category_D, category_A_avg, category_D_avg
  
def write_excel(args, table, threshold, top_K_category_list, interest_group, interest_GA, interest_GD, GA_all_avg, GD_all_avg, GA_extreme_avg, GD_extreme_avg, all_user_num, extreme_user_num, target_GA_avg, target_GD_avg, target_extreme_GA_avg, target_extreme_GD_avg, row_offset, column_offset):
    
    table.cell(row = 1 + row_offset, column = 1 + column_offset, value = args.model +'_' + 'top_' + str(eval(args.topN)[0]) + '_'+ str(threshold) + '_alpha=' + str(args.alpha))
    table.cell(row = 1 + row_offset, column = 2 + column_offset, value = "Interest")
    table.cell(row = 2 + row_offset, column = 1 + column_offset, value = "Top_K_category")
    table.cell(row = 1 + row_offset, column = 3 + column_offset, value = "All user")
    table.cell(row = 1 + row_offset, column = 4 + column_offset, value = "Extreme user")
    column = 5 + column_offset
    for interest in interest_group:
        table.cell(row = 1 + row_offset, column = column, value = interest)
        table.cell(row = 3 + row_offset, column = column, value = len(interest_group[interest]))
        column += 1 
    table.cell(row = 3 + row_offset, column = 2 + column_offset, value = 'size')
    table.cell(row = 3 + row_offset, column = 3 + column_offset, value = all_user_num[threshold])
    table.cell(row = 3 + row_offset, column = 4 + column_offset, value = extreme_user_num[threshold])


          
    row = 4 + row_offset
    for top_K_category in top_K_category_list:
        table.cell(row = row, column = 1 + column_offset, value = top_K_category)
        table.cell(row = row, column = 2 + column_offset, value = 'GA')
        table.cell(row = row, column = 3 + column_offset, value = GA_all_avg[threshold][top_K_category])
        table.cell(row = row, column = 4 + column_offset, value = GA_extreme_avg[threshold][top_K_category])


        column = 5 + column_offset
        for interest in interest_group:
            table.cell(row = row, column = column, value = interest_GA[threshold][top_K_category][interest])
            column += 1
        row += 1
    
    table.cell(row = row, column = 1 + column_offset, value = 'target')
    table.cell(row = row, column = 2 + column_offset, value = 'GA')
    table.cell(row = row, column = 3 + column_offset, value = target_GA_avg[threshold])
    table.cell(row = row, column = 4 + column_offset, value = target_extreme_GA_avg[threshold])
         
    row = 5+len(top_K_category_list) + row_offset
    for top_K_category in top_K_category_list:
        table.cell(row = row, column = 2 + column_offset, value = 'GD')
        table.cell(row = row, column = 1 + column_offset, value = top_K_category)
        table.cell(row = row, column = 3 + column_offset, value = GD_all_avg[threshold][top_K_category])
        table.cell(row = row, column = 4 + column_offset, value = GD_extreme_avg[threshold][top_K_category])
        column = 5 + column_offset
        for interest in interest_group:
            table.cell(row = row, column = column, value = interest_GD[threshold][top_K_category][interest])
            column += 1
        row += 1
       
    table.cell(row = row, column = 1 + column_offset, value = 'target')
    table.cell(row = row, column = 2 + column_offset, value = 'GD')
    table.cell(row = row, column = 3 + column_offset, value = target_GD_avg[threshold])
    table.cell(row = row, column = 4 + column_offset, value = target_extreme_GD_avg[threshold])
     
def write_excel_category(table, threshold_category_list, category_A_avg, category_D_avg, row_offset, column_offset):
    table.cell(row = 1 + row_offset, column = 1 + column_offset, value = 'category_size_threshold')
    table.cell(row = 1 + row_offset, column = 2 + column_offset, value = 'category_A_avg')
    table.cell(row = 1 + row_offset, column = 3 + column_offset, value = 'category_D_avg')
    
    row = 2 + row_offset
    for threshold_category in threshold_category_list:
        table.cell(row = row, column = 1 + column_offset, value = threshold_category)
        row += 1
        
    row = 2 + row_offset
    for threshold_category in threshold_category_list:
        table.cell(row = row, column = 2 + column_offset, value = category_A_avg[threshold_category])
        table.cell(row = row, column = 3 + column_offset, value = category_D_avg[threshold_category])
        row += 1

def write_excel_all(args, excel_file, threshold_proportion_list, threshold_group_number, threshold_category_list, top_K_category_list, train_dict, item_category, category_list, interest_GA, interest_GD, GA_all_avg, GD_all_avg, GA_extreme_avg, GD_extreme_avg, all_user_num, extreme_user_num, target_GA_avg, target_GD_avg, target_extreme_GA_avg, target_extreme_GD_avg, category_A_avg, category_D_avg):
    try:
        load_workbook(filename=excel_file)
    except:
        workbook = openpyxl.Workbook()
    else:
        workbook = load_workbook(filename=excel_file)

    table = workbook.active
    column_offset = 0
    row_offset = 0

    # if args.model == 'FM':
    #     column_offset = 0
    # if args.model == 'NFM':
    #     column_offset = 14
        
    if args.alpha == 0:
        column_offset = 0

    if args.alpha == 1:
        column_offset += 14

    if args.alpha == 2:
        column_offset += 28

    if args.alpha == 100:
        column_offset += 42

    if eval(args.topN)[0] == 10:
        row_offset = 0

    # if top_N == 20:
    #     row_offset = len(threshold_proportion_list)*(len(top_K_category_list) *2 +9) 

    # if top_N == 50:
    #     row_offset = 2*len(threshold_proportion_list)*(len(top_K_category_list) *2 +9) 
    if args.model == 'random':
        row_offset = 2*len(threshold_proportion_list)*(len(top_K_category_list) *2 +9) 
        column_offset = 0

    for threshold in threshold_proportion_list:
        interest_group = get_interest_group(train_dict, item_category, category_list, threshold, threshold_group_number)

        # for top_K_category in top_K_category_list:
        write_excel(args, table, threshold, top_K_category_list, interest_group, interest_GA, interest_GD, GA_all_avg, GD_all_avg, GA_extreme_avg, GD_extreme_avg, all_user_num, extreme_user_num, target_GA_avg, target_GD_avg, target_extreme_GA_avg, target_extreme_GD_avg, row_offset, column_offset)
        row_offset += len(top_K_category_list) *2 +5 


        write_excel_category(table, threshold_category_list, category_A_avg, category_D_avg, row_offset, column_offset)

        row_offset += 4
    workbook.save(excel_file)

   
    